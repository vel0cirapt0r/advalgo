import pandas as pd
import numpy as np
from collections import defaultdict
import time
from openpyxl import Workbook, load_workbook
import logging
import os
from datetime import datetime
import colorama
from colorama import Fore, Style
import argparse
import random

class ColoredFormatter(logging.Formatter):
    """Custom formatter to add colors to log messages based on their level."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.color_map = {
            logging.DEBUG: Fore.BLUE,
            logging.INFO: Fore.GREEN,
            logging.WARNING: Fore.YELLOW,
            logging.ERROR: Fore.RED,
            logging.CRITICAL: Fore.RED + Style.BRIGHT,
        }

    def format(self, record):
        s = super().format(record)
        color = self.color_map.get(record.levelno, '')
        return color + s + Style.RESET_ALL


def setup_logging(log_file=f'log/{datetime.now()}.log', level=logging.DEBUG):
    """Configure the root logger with file and console handlers."""
    # Initialize Colorama for cross-platform color support
    colorama.init()

    # Create log directory if it doesn't exist
    log_dir = os.path.dirname(log_file)
    if log_dir and not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # Set up file handler with standard formatter
    file_handler = logging.FileHandler(log_file)
    file_formatter = logging.Formatter(
        '%(asctime)s.%(msecs)03d [%(levelname)s] %(name)s.%(funcName)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(file_formatter)
    file_handler.setLevel(logging.DEBUG)  # Always log DEBUG and above to file

    # Set up console handler with colored formatter
    console_handler = logging.StreamHandler()
    console_formatter = ColoredFormatter(
        '%(asctime)s.%(msecs)03d [%(levelname)s] %(name)s.%(funcName)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_handler.setFormatter(console_formatter)

    # Set console handler level based on DEBUG_MODE in constants.py
    try:
        from constants import DEBUG_MODE
        if DEBUG_MODE:
            console_handler.setLevel(logging.DEBUG)
        else:
            console_handler.setLevel(logging.INFO)
    except (ImportError, AttributeError):
        # Default to INFO if DEBUG_MODE is not defined
        console_handler.setLevel(logging.INFO)

    # Configure the root logger
    root_logger = logging.getLogger()
    if not root_logger.handlers:  # Avoid duplicate handlers
        root_logger.addHandler(file_handler)
        root_logger.addHandler(console_handler)
        root_logger.setLevel(level)

def load_data(file_path, strict_order_check=False):
    df = pd.read_csv(file_path)
    df['order_id'] = df['order_id'].astype('int32')
    df['product_id'] = df['product_id'].astype('int32')
    df['add_to_cart_order'] = df['add_to_cart_order'].astype('int32')
    df['reordered'] = df['reordered'].fillna(0).astype('int32')

    df = df.sort_values(['order_id', 'add_to_cart_order'])

    if strict_order_check:
        def validate_order(group):
            orders = group['add_to_cart_order'].values
            if not np.all(np.diff(orders) > 0):
                logging.warning(f"Non-increasing add_to_cart_order in order_id {group.name}")
        df.groupby('order_id').apply(validate_order)

    sequences = df.groupby('order_id')['product_id'].apply(list).to_dict()
    return df, sequences


# Build model from training prefixes
def build_model(train_df, alpha=0.1, top_n=100, window=5, reordered_boost=1.15):
    unique_products = train_df['product_id'].unique()
    V_train = len(unique_products)
    num_orders_train = train_df['order_id'].nunique()
    avg_seq_len_train = train_df.groupby('order_id').size().mean()

    # Estimate cost
    est_cost = num_orders_train * avg_seq_len_train * (avg_seq_len_train - 1) / 2
    if est_cost > 1e6:
        logging.warning(f"High co-vis cost estimated: {est_cost:.2e} operations (orders={num_orders_train}, avg_len={avg_seq_len_train:.2f})")

    # Markov transitions
    pairs = train_df.assign(next_product=train_df.groupby('order_id')['product_id'].shift(-1)).dropna(
        subset=['next_product'])
    pairs['next_product'] = pairs['next_product'].astype('int32')
    trans_counts = pairs.groupby(['product_id', 'next_product']).size()

    transitions = defaultdict(dict)
    for (current, next_), count in trans_counts.items():
        transitions[current][next_] = count

    for current in list(transitions):
        # Prune to top 100 by count first
        if len(transitions[current]) > top_n:
            sorted_neighbors = sorted(transitions[current].items(), key=lambda x: x[1], reverse=True)[:top_n]
            transitions[current] = dict(sorted_neighbors)

        # Now normalize with smoothing
        seen = len(transitions[current])
        total = sum(transitions[current].values())
        denom = total + alpha * seen
        for next_ in list(transitions[current]):
            transitions[current][next_] = (transitions[current][next_] + alpha) / denom

    # Co-visitation (symmetric within window)
    co_vis_counts = defaultdict(lambda: defaultdict(float))
    order_groups = train_df.groupby('order_id')
    num_groups = train_df['order_id'].nunique()
    processed = 0
    for name, g in order_groups:
        products = g['product_id'].values
        reords = g['reordered'].values
        n = len(products)
        for i in range(n):
            for j in range(i + 1, min(i + window + 1, n)):
                dist = j - i
                weight = 1.0 / dist
                # Forward boost if target reordered > 0 (NaN filled as 0 in load_data)
                boost_fwd = reordered_boost if reords[j] > 0 else 1.0
                co_vis_counts[products[i]][products[j]] += weight * boost_fwd
                # Reverse boost if target reordered > 0
                boost_rev = reordered_boost if reords[i] > 0 else 1.0
                co_vis_counts[products[j]][products[i]] += weight * boost_rev
        processed += 1
        if processed % 1000 == 0:
            logging.debug(f"Processed {processed}/{num_groups} orders for co-vis")

    co_vis = defaultdict(dict)
    for current in list(co_vis_counts):
        # Prune to top 100 by count first
        if len(co_vis_counts[current]) > top_n:
            sorted_neighbors = sorted(co_vis_counts[current].items(), key=lambda x: x[1], reverse=True)[:top_n]
            co_vis_counts[current] = dict(sorted_neighbors)

        # Normalize with smoothing
        seen = len(co_vis_counts[current])
        total = sum(co_vis_counts[current].values())
        denom = total + alpha * seen
        for next_ in list(co_vis_counts[current]):
            co_vis[current][next_] = (co_vis_counts[current][next_] + alpha) / denom

    # Popularity from last items in prefixes
    last_items = train_df.groupby('order_id')['product_id'].last()
    pop_counts = last_items.value_counts()
    total_pop = pop_counts.sum()
    popularity = {item: count / total_pop for item, count in pop_counts.items()} if total_pop > 0 else {}

    return transitions, co_vis, popularity, V_train, num_orders_train, avg_seq_len_train


# Predict next item
def predict_next(transitions, co_vis, popularity, sequence, tail_weights, blend_markov, blend_cov):
    if len(sequence) == 0:
        if not popularity:
            return None, 0, 0.0, 'pop'
        pred = max(popularity, key=popularity.get)
        return pred, 0, popularity.get(pred, 0.0), 'pop'

    tail = sequence[-3:]  # oldest to newest
    weights = tail_weights[-len(tail):]  # weights correspond to tail order: lower for older, higher for newer

    candidates = set()
    markov_scores = defaultdict(float)
    co_scores = defaultdict(float)
    for i, t in enumerate(tail):
        w = weights[i]
        if t in transitions and transitions[t]:
            for j, p in transitions[t].items():
                markov_scores[j] += w * p
        if t in co_vis and co_vis[t]:
            for j, p in co_vis[t].items():
                co_scores[j] += w * p

    candidates = set(list(markov_scores.keys()) + list(co_scores.keys()))

    if not candidates:
        if not popularity:
            return None, 0, 0.0, 'pop'
        pred = max(popularity, key=popularity.get)
        return pred, 0, popularity.get(pred, 0.0), 'pop'

    # No exclusion of seen items to allow repeats

    if not candidates:
        if not popularity:
            return None, 0, 0.0, 'pop'
        pred = max(popularity, key=popularity.get)
        return pred, 0, popularity.get(pred, 0.0), 'pop'

    # Blend scores
    scores = {}
    for j in candidates:
        s_markov = markov_scores.get(j, 0.0)
        s_co = co_scores.get(j, 0.0)
        scores[j] = blend_markov * s_markov + blend_cov * s_co

    # For determinism: sort by blended score desc, then markov desc, then popularity desc, then product_id asc
    sorted_candidates = sorted(scores.items(),
                               key=lambda x: (-x[1], -markov_scores.get(x[0], 0), -popularity.get(x[0], 0), x[0]))
    predicted = sorted_candidates[0][0]
    pred_score = scores[predicted]
    num_cand = len(candidates)

    s_markov = markov_scores.get(predicted, 0.0)
    s_co = co_scores.get(predicted, 0.0)
    if s_markov > s_co:
        dominant = 'markov'
    elif s_co > s_markov:
        dominant = 'cov'
    else:
        dominant = 'tie'

    return predicted, num_cand, pred_score, dominant


# Evaluate for a given k
def evaluate(sequences, transitions, co_vis, popularity, k, tail_weights, blend_markov, blend_cov):
    correct = 0
    total = 0
    num_backoff = 0
    count_markov = 0
    count_cov = 0
    count_tie = 0
    predictions = []

    for order_id, seq in sequences.items():
        if len(seq) > k:
            input_seq = seq[:-k]
            true_last = seq[-1]
            predicted, num_cand, pred_score, dominant = predict_next(transitions, co_vis, popularity, input_seq, tail_weights,
                                                             blend_markov, blend_cov)
            if predicted is None:
                continue
            used_backoff = num_cand == 0
            if used_backoff:
                num_backoff += 1
                dominant = 'pop'
            else:
                if dominant == 'markov':
                    count_markov += 1
                elif dominant == 'cov':
                    count_cov += 1
                elif dominant == 'tie':
                    count_tie += 1
            if predicted == true_last:
                correct += 1
            backoff_score = pred_score if used_backoff else None
            pop_rank = None
            if used_backoff and popularity:
                pop_items = sorted(popularity, key=popularity.get, reverse=True)
                pop_rank = pop_items.index(predicted) + 1 if predicted in pop_items else None
            predictions.append({
                'order_id': order_id,
                'k': k,
                'input_sequence': input_seq,
                'true_last': true_last,
                'predicted': predicted,
                'correct': predicted == true_last,
                'used_backoff': used_backoff,
                'num_candidates': num_cand,
                'score_of_prediction': pred_score,
                'backoff_score': backoff_score,
                'pop_rank_if_backoff': pop_rank,
                'dominant_model': dominant
            })
            total += 1

    accuracy = correct / total if total > 0 else 0
    non_backoff = total - num_backoff if total > 0 else 0
    dominance_markov = 100 * count_markov / non_backoff if non_backoff > 0 else 0
    dominance_cov = 100 * count_cov / non_backoff if non_backoff > 0 else 0
    dominance_tie = 100 * count_tie / non_backoff if non_backoff > 0 else 0
    return accuracy, predictions, total, num_backoff, correct, dominance_markov, dominance_cov, dominance_tie


# Main function
def main(input_csv, output_excel, alpha, window, top_n, tail_weights, blend_markov, blend_cov, reordered_boost, seed, strict_order_check, max_orders, sample_rate, no_repeats):
    if len(tail_weights) < 1:
        raise ValueError("tail_weights must have at least 1 element")
    # Note: tail_weights mapped oldest (low weight) to newest (high weight), using last min(3, len(tail_weights))

    np.random.seed(seed)
    random.seed(seed)
    pd.set_option('mode.chained_assignment', 'warn')

    log_file = f'log/{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.log'
    setup_logging(log_file=log_file)

    params_str = f"alpha={alpha}, window={window}, top_n={top_n}, tail_weights={tail_weights}, blend_markov={blend_markov}, blend_cov={blend_cov}, reordered_boost={reordered_boost}, seed={seed}, max_orders={max_orders}, sample_rate={sample_rate}, no_repeats={no_repeats}"
    logging.info(f"Starting run on {input_csv}, params: {params_str}")

    start_load = time.time()
    df, sequences = load_data(input_csv, strict_order_check)
    load_time = time.time() - start_load

    logging.info(f"Dataset size: {len(df)} rows, {len(sequences)} orders")

    results = {}
    all_predictions = []
    total_train_time = 0
    total_eval_time = 0
    total_correct = 0
    total_eligible = 0
    metrics_per_k = {}
    last_popularity = None
    last_transitions = None
    last_k = None

    for k in [1, 2, 3]:
        train_start = time.time()
        order_ids = list(df['order_id'].unique())
        eligible_ids = [oid for oid in order_ids if len(sequences[oid]) > k]
        eligible = len(eligible_ids)
        ineligible = len(sequences) - eligible
        if eligible == 0:
            continue
        sampled_ids = eligible_ids
        if sample_rate < 1.0:
            sample_size = int(eligible * sample_rate)
            sampled_ids = random.sample(eligible_ids, sample_size)
        train_groups = []
        order_groups = df[df['order_id'].isin(sampled_ids)].groupby('order_id')
        for name, g in order_groups:
            train_groups.append(g.iloc[:-k])
            if max_orders and len(train_groups) >= max_orders:
                break
        train_df = pd.concat(train_groups)
        transitions, co_vis, popularity, V_train, num_orders_train, avg_seq_len_train = build_model(train_df, alpha,
                                                                                                    top_n, window, reordered_boost)
        train_time = time.time() - train_start
        total_train_time += train_time

        # Sanity checks
        fail_count = 0
        sample_curr = random.sample(list(transitions.keys()), min(100, len(transitions)))
        for curr in sample_curr:
            prob_sum = sum(transitions[curr].values())
            if not 0.99 < prob_sum < 1.01:
                logging.warning(f"Prob sum {prob_sum} for {curr}")
                fail_count += 1
        sample_curr = random.sample(list(co_vis.keys()), min(100, len(co_vis)))
        for curr in sample_curr:
            prob_sum = sum(co_vis[curr].values())
            if not 0.99 < prob_sum < 1.01:
                logging.warning(f"Prob sum {prob_sum} for {curr}")
                fail_count += 1
        logging.info(f"Sanity check: {fail_count} failures out of 200 samples")

        eval_start = time.time()
        acc, preds, _, num_backoff, correct_k, dom_markov, dom_cov, dom_tie = evaluate(sequences, transitions, co_vis, popularity, k, tail_weights,
                                                         blend_markov, blend_cov)
        eval_time = time.time() - eval_start
        total_eval_time += eval_time

        results[f'Accuracy for hide last {k}'] = acc
        coverage_model = 100 * (eligible - num_backoff) / eligible if eligible > 0 else 0
        logging.debug(f"For k={k}:")
        logging.debug(f"  Number of eligible users: {eligible}")
        logging.debug(f"  Ineligible users (len <= k): {ineligible}")
        logging.debug(f"  Coverage (% predictions via Model): {coverage_model:.2f}%")
        logging.debug(f"  Train time: {train_time:.2f}s")
        logging.debug(f"  Eval time: {eval_time:.2f}s")

        metrics_per_k[k] = {
            'eligible_users': eligible,
            'ineligible_users (len <= k)': ineligible,
            'correct_predictions': correct_k,
            'num_backoff': num_backoff,
            'coverage_model_%': coverage_model,
            'coverage_definition': '% of eligible eval users not using popularity backoff',
            'train_time_s': train_time,
            'eval_time_s': eval_time,
            'num_products_train': V_train,
            'num_orders_train': num_orders_train,
            'avg_seq_len_train': avg_seq_len_train,
            'topN_pruning': top_n,
            'alpha': alpha,
            'blend_markov': blend_markov,
            'blend_cov': blend_cov,
            'tail_weights': str(tail_weights),
            'window': window,
            'reordered_boost': reordered_boost,
            'dominance_markov_%': dom_markov,
            'dominance_cov_%': dom_cov,
            'dominance_tie_%': dom_tie,
            'dominance_definition': '% of non-backoff predictions where Markov score > Co-vis score (markov), etc.'
        }

        all_predictions.extend(preds)
        total_correct += correct_k
        total_eligible += eligible

        last_popularity = popularity
        last_transitions = transitions
        last_k = k

    macro_acc = np.mean(list(results.values()))
    micro_acc = total_correct / total_eligible if total_eligible > 0 else 0
    results['Overall accuracy (macro – mean of per-k)'] = macro_acc
    results['Overall accuracy (micro – global)'] = micro_acc

    # Write to Excel
    write_start = time.time()
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    ws_summary.append(['Metric', 'Value'])
    ws_summary.append(['Run timestamp', datetime.now()])
    ws_summary.append(['Input file', input_csv])
    ws_summary.append(['Dataset rows', len(df)])
    ws_summary.append(['Total orders', len(sequences)])
    for key, val in results.items():
        ws_summary.append([key, val])
    for k in [1,2,3]:
        if k in metrics_per_k:
            ws_summary.append([f'Coverage for k={k}', metrics_per_k[k]['coverage_model_%']])
            ws_summary.append([f'Dominance Markov for k={k}', metrics_per_k[k]['dominance_markov_%']])
            ws_summary.append([f'Dominance Co-vis for k={k}', metrics_per_k[k]['dominance_cov_%']])
            ws_summary.append([f'Dominance Tie for k={k}', metrics_per_k[k]['dominance_tie_%']])
    ws_summary.append(['Load time', load_time])
    ws_summary.append(['Total train time', total_train_time])
    ws_summary.append(['Total eval time', total_eval_time])
    ws_summary.append(['Write time', ''])
    ws_summary.append(['Seed', seed])
    ws_summary.append(['Note on stochasticity', 'Minimal (only affects sampling if sample_rate<1)'])
    ws_summary.append(['Note on pruning', 'Normalized after pruning topN'])
    ws_summary.append(['Note on coverage', '% of eligible eval users not using popularity backoff (eval on all, train may be sampled)'])

    ws_summary.append([])
    ws_summary.append(['Params', ''])
    ws_summary.append(['alpha', alpha])
    ws_summary.append(['window', window])
    ws_summary.append(['topN_neighbors', top_n])
    ws_summary.append(['tail_weights', str(tail_weights)])
    ws_summary.append(['blend', f'{blend_markov} Markov / {blend_cov} Co-vis'])
    ws_summary.append(['reordered_boost', reordered_boost])

    ws_preds = wb.create_sheet('Predictions')
    headers = ['order_id', 'k', 'input_sequence', 'true_last', 'predicted', 'correct', 'used_backoff', 'num_candidates',
               'score_of_prediction', 'backoff_score', 'pop_rank_if_backoff', 'dominant_model']
    ws_preds.append(headers)
    for p in all_predictions:
        input_str = ','.join(map(str, p['input_sequence']))
        ws_preds.append(
            [p['order_id'], p['k'], input_str, p['true_last'], p['predicted'], p['correct'], p['used_backoff'],
             p['num_candidates'], p['score_of_prediction'], p['backoff_score'], p['pop_rank_if_backoff'], p['dominant_model']])

    for k in [1, 2, 3]:
        if k in metrics_per_k:
            ws_metrics = wb.create_sheet(f'Metrics_k={k}')
            ws_metrics.append(['Metric', 'Value'])
            for key, val in metrics_per_k[k].items():
                ws_metrics.append([key, val])

    if all_predictions:
        df_preds = pd.DataFrame(all_predictions)
        errors = df_preds[~df_preds['correct']]
        top_errors = errors.groupby(['true_last', 'predicted']).size().reset_index(name='count').sort_values('count', ascending=False).head(100)
        ws_errors = wb.create_sheet('Top_Errors')
        ws_errors.append(['true_last', 'predicted', 'count'])
        for _, row in top_errors.iterrows():
            ws_errors.append([row['true_last'], row['predicted'], row['count']])

    if last_popularity and last_transitions:
        pop_top20 = sorted(last_popularity, key=last_popularity.get, reverse=True)[:20]
        ws_trans = wb.create_sheet('Model_TopTransitions')
        ws_trans.append(['Note', f'From k={last_k} model'])
        ws_trans.append(['product', 'neighbor', 'prob'])
        for p in pop_top20:
            if p in last_transitions:
                top_neigh = sorted(last_transitions[p].items(), key=lambda x: x[1], reverse=True)[:5]
                for n, prob in top_neigh:
                    ws_trans.append([p, n, prob])

    wb.save(output_excel)
    write_time = time.time() - write_start

    # Update write time in summary
    wb = load_workbook(output_excel)
    ws_summary = wb['Summary']
    for row in ws_summary.iter_rows():
        if row[0].value == 'Write time':
            row[1].value = write_time
            break
    wb.save(output_excel)

    logging.info(f"Load time: {load_time:.2f}s")
    logging.info(f"Total train time: {total_train_time:.2f}s")
    logging.info(f"Total eval time: {total_eval_time:.2f}s")
    logging.info(f"Write time: {write_time:.2f}s")
    logging.info(f"Results saved to {output_excel}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input_csv", nargs='?', default="order-Product_prior.csv")
    parser.add_argument("output_excel", nargs='?', default="results.xlsx")
    parser.add_argument('--alpha', type=float, default=0.1)
    parser.add_argument('--window', type=int, default=5)
    parser.add_argument('--top_n', type=int, default=100)
    parser.add_argument('--tail_weights', nargs='+', type=float, default=[0.3, 0.6, 1.0])
    parser.add_argument('--blend_markov', type=float, default=0.6)
    parser.add_argument('--blend_cov', type=float, default=0.4)
    parser.add_argument('--reordered_boost', type=float, default=1.15)
    parser.add_argument('--seed', type=int, default=42)
    parser.add_argument('--strict_order_check', action='store_true')
    parser.add_argument('--max_orders', type=int, default=None)
    parser.add_argument('--sample_rate', type=float, default=1.0)
    parser.add_argument('--no_repeats', action='store_true')
    args = parser.parse_args()
    main(args.input_csv, args.output_excel, args.alpha, args.window, args.top_n, args.tail_weights, args.blend_markov, args.blend_cov, args.reordered_boost, args.seed, args.strict_order_check, args.max_orders, args.sample_rate, args.no_repeats)
